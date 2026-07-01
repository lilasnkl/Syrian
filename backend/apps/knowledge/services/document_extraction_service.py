import csv
import hashlib
import io
import re

from django.core.files.storage import default_storage

from apps.knowledge.models import KnowledgeDocument
from shared.exceptions import ExternalServiceError


class DocumentExtractionService:
    EXTRACTOR_VERSION = "v1"

    @classmethod
    def extract(cls, source) -> KnowledgeDocument:
        raw_bytes = cls._read_source_bytes(source)
        text, metadata = cls._extract_text(raw_bytes, source=source)
        normalized = cls._normalize_text(text)
        text_hash = hashlib.sha256(normalized.encode("utf-8")).hexdigest() if normalized else ""

        return KnowledgeDocument.objects.create(
            source=source,
            provider=source.provider,
            extraction_status=KnowledgeDocument.STATUS_EXTRACTED,
            raw_text=normalized,
            normalized_text_hash=text_hash,
            detected_language=cls._detect_language(normalized),
            page_count=metadata.get("page_count", 0),
            row_count=metadata.get("row_count", 0),
            token_count=cls._count_tokens(normalized),
            extractor_version=cls.EXTRACTOR_VERSION,
        )

    @staticmethod
    def _read_source_bytes(source) -> bytes:
        if not source.storage_path:
            return source.description.encode("utf-8")
        with default_storage.open(source.storage_path, "rb") as file_handle:
            return file_handle.read()

    @classmethod
    def _extract_text(cls, raw_bytes: bytes, *, source) -> tuple[str, dict]:
        content_type = source.content_type or ""
        filename = (source.original_filename or "").lower()

        if content_type in {"text/plain", "text/markdown"} or filename.endswith((".txt", ".md")):
            return cls._decode_text(raw_bytes), {}
        if content_type in {"text/csv", "application/csv"} or filename.endswith(".csv"):
            return cls._extract_csv(raw_bytes), {}
        if content_type == "application/pdf" or filename.endswith(".pdf"):
            return cls._extract_pdf(raw_bytes)
        if content_type.endswith("spreadsheetml.sheet") or filename.endswith(".xlsx"):
            return cls._extract_xlsx(raw_bytes)
        if content_type.endswith("wordprocessingml.document") or filename.endswith(".docx"):
            return cls._extract_docx(raw_bytes)

        raise ExternalServiceError(
            detail="Unsupported document format.",
            code="knowledge_extraction_unsupported_format",
            details={"content_type": content_type, "filename": source.original_filename},
        )

    @staticmethod
    def _decode_text(raw_bytes: bytes) -> str:
        for encoding in ("utf-8-sig", "utf-8", "cp1256", "latin-1"):
            try:
                return raw_bytes.decode(encoding)
            except UnicodeDecodeError:
                continue
        return raw_bytes.decode("utf-8", errors="ignore")

    @classmethod
    def _extract_csv(cls, raw_bytes: bytes) -> str:
        decoded = cls._decode_text(raw_bytes)
        reader = csv.reader(io.StringIO(decoded))
        lines = []
        for index, row in enumerate(reader, start=1):
            values = [value.strip() for value in row if value and value.strip()]
            if values:
                lines.append(f"Row {index}: " + " | ".join(values))
        return "\n".join(lines)

    @staticmethod
    def _extract_pdf(raw_bytes: bytes) -> tuple[str, dict]:
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise ExternalServiceError(
                detail="PDF extraction dependency is not installed.",
                code="pdf_extractor_missing",
                details={"detail": "Install pypdf."},
            ) from exc

        reader = PdfReader(io.BytesIO(raw_bytes))
        pages = []
        for index, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            if text.strip():
                pages.append(f"Page {index}\n{text}")
        return "\n\n".join(pages), {"page_count": len(reader.pages)}

    @staticmethod
    def _extract_xlsx(raw_bytes: bytes) -> tuple[str, dict]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ExternalServiceError(
                detail="Spreadsheet extraction dependency is not installed.",
                code="xlsx_extractor_missing",
                details={"detail": "Install openpyxl."},
            ) from exc

        workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
        lines = []
        row_count = 0
        for worksheet in workbook.worksheets:
            lines.append(f"Sheet: {worksheet.title}")
            for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                values = [str(value).strip() for value in row if value not in {None, ""}]
                if values:
                    row_count += 1
                    lines.append(f"Row {row_index}: " + " | ".join(values))
        return "\n".join(lines), {"row_count": row_count}

    @staticmethod
    def _extract_docx(raw_bytes: bytes) -> tuple[str, dict]:
        try:
            from docx import Document
        except ImportError as exc:
            raise ExternalServiceError(
                detail="DOCX extraction dependency is not installed.",
                code="docx_extractor_missing",
                details={"detail": "Install python-docx."},
            ) from exc

        document = Document(io.BytesIO(raw_bytes))
        paragraphs = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]
        return "\n".join(paragraphs), {}

    @staticmethod
    def _normalize_text(text: str) -> str:
        normalized_lines = [re.sub(r"\s+", " ", line).strip() for line in (text or "").splitlines()]
        return "\n".join(line for line in normalized_lines if line)

    @staticmethod
    def _detect_language(text: str) -> str:
        if re.search(r"[\u0600-\u06FF]", text or ""):
            return "ar"
        return "en" if re.search(r"[A-Za-z]", text or "") else ""

    @staticmethod
    def _count_tokens(text: str) -> int:
        return len((text or "").split())

